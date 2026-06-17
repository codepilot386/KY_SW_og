"""
DB → 엑셀(.xlsx) 내보내기
─────────────────────────
회원(users)과 훈련기록(history)을 엑셀 파일 하나로 내보낸다.

사용법:
    venv\\Scripts\\python.exe export_db.py

결과:
    backend/export/훈련앱_데이터_YYYYMMDD_HHMMSS.xlsx  (시트 2개: 회원 / 훈련기록)
"""
import sqlite3, os, json, time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(BASE_DIR, "app.db")
OUT_DIR  = os.path.join(BASE_DIR, "export")

HEADER_FILL = PatternFill("solid", fgColor="3182F6")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _autosize_and_style(ws, ncols):
    """헤더 스타일 + 컬럼 너비 자동 조정"""
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, 10), 60)
    ws.freeze_panes = "A2"


def export():
    if not os.path.exists(DB_PATH):
        print(f"❌ DB 파일이 없습니다: {DB_PATH}")
        print("   먼저 앱을 실행해 회원가입을 한 번 진행하면 DB가 생성됩니다.")
        return

    os.makedirs(OUT_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    wb = Workbook()

    # ── 시트 1: 회원 ──────────────────────────────
    ws_u = wb.active
    ws_u.title = "회원"
    u_headers = ["회원번호", "아이디", "가입일시"]
    ws_u.append(u_headers)
    for r in conn.execute("SELECT id, username, created_at FROM users ORDER BY id"):
        ws_u.append([r["id"], r["username"], r["created_at"]])
    _autosize_and_style(ws_u, len(u_headers))

    # ── 시트 2: 훈련기록 ──────────────────────────
    ws_h = wb.create_sheet("훈련기록")
    h_headers = ["기록번호", "회원아이디", "시나리오", "난이도", "등급", "점수",
                 "취약도(%)", "위험등급", "행동유형", "총평", "핵심팁", "훈련일시"]
    ws_h.append(h_headers)

    sql = """
        SELECT h.id, u.username, h.scenario, h.difficulty, h.grade, h.score,
               h.feedback, h.tip, h.report_json, h.created_at
        FROM history h LEFT JOIN users u ON u.id = h.user_id
        ORDER BY h.id
    """
    for r in conn.execute(sql):
        report = {}
        if r["report_json"]:
            try: report = json.loads(r["report_json"])
            except Exception: report = {}
        ws_h.append([
            r["id"], r["username"], r["scenario"], r["difficulty"], r["grade"], r["score"],
            report.get("susceptibility_percent", ""),
            report.get("risk_level", ""),
            report.get("user_type", ""),
            r["feedback"] or report.get("summary", ""),
            r["tip"] or report.get("tip", ""),
            r["created_at"],
        ])
    _autosize_and_style(ws_h, len(h_headers))

    conn.close()

    stamp = time.strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(OUT_DIR, f"훈련앱_데이터_{stamp}.xlsx")
    wb.save(out_path)

    n_users = ws_u.max_row - 1
    n_hist  = ws_h.max_row - 1
    print(f"✅ 엑셀 저장 완료: {out_path}")
    print(f"   - 회원 {n_users}명 / 훈련기록 {n_hist}건")
    print(f"   더블클릭하면 엑셀에서 바로 열립니다.")


if __name__ == "__main__":
    export()
